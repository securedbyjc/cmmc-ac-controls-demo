#!/usr/bin/env python3
"""
CMMC Level 2 AC Controls Assessment Demo
Educational Version for Aspire Cyber Podcast
Demonstrates automated evidence generation for Access Control requirements
"""

import argparse
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np  # pyright: ignore[reportMissingImports]
import pandas as pd  # pyright: ignore[reportMissingModuleSource]
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side  # pyright: ignore[reportMissingModuleSource]
from openpyxl.utils import get_column_letter  # type: ignore

print("=" * 60)
print("🎯 CMMC LEVEL 2 AC CONTROLS ASSESSMENT DEMO")
print("📺 Aspire Cyber Podcast Educational Version")
print("=" * 60)

# Mock company configuration
MOCK_COMPANY = {
    "name": "TechDefense Solutions LLC",
    "type": "Defense Subcontractor",
    "employees": 250,
    "cui_users": 85,
    "environment": "Hybrid (GCP + On-premise)",
    "assessment_date": datetime.now().strftime("%Y-%m-%d"),
}

# AC Controls configuration
AC_CONTROLS = {
    "AC.L2-3.1.1": {
        "name": "Limit System Access",
        "description": "Limit system access to authorized users, processes, and devices",
        "evidence_types": ["Access Matrix", "User List", "Access Reviews"],
        "test_method": "Review access control matrix and sample user permissions",
    },
    "AC.L2-3.1.3": {
        "name": "Control CUI Flow",
        "description": "Control the flow of CUI in accordance with approved authorizations",
        "evidence_types": ["Network Diagram", "DLP Policies", "VPC Config"],
        "test_method": "Review network segmentation and data flow controls",
    },
    "AC.L2-3.1.6": {
        "name": "Session Lock",
        "description": "Use session lock with pattern-hiding displays",
        "evidence_types": ["GPO Settings", "Workspace Policy", "Test Results"],
        "test_method": "Verify timeout settings and test sample workstations",
    },
    "AC.L2-3.1.7": {
        "name": "Prevent Identifier Reuse",
        "description": "Prevent reuse of identifiers for a defined period",
        "evidence_types": ["Identity Policy", "Archived Users", "Retention Logs"],
        "test_method": "Review identity lifecycle management and retention",
    },
    "AC.L2-3.1.12": {
        "name": "Monitor Remote Access",
        "description": "Monitor and control remote access sessions",
        "evidence_types": ["VPN Logs", "Session Recordings", "SIEM Alerts"],
        "test_method": "Review remote access logs and monitoring configuration",
    },
    "AC.L2-3.1.20": {
        "name": "Control External Connections",
        "description": "Control and monitor external system connections",
        "evidence_types": ["Connection Inventory", "Monitoring Dashboard", "Attestations"],
        "test_method": "Review external connections and continuous monitoring",
    },
}


def generate_ac_evidence():
    """Generate mock evidence for AC controls with multiple rows for multiple evidence"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    evidence = [
        # AC.L2-3.1.1 - First evidence
        {
            "control_id": "AC.L2-3.1.1",
            "control_name": AC_CONTROLS["AC.L2-3.1.1"]["name"],
            "status": "COMPLIANT",
            "implementation": "Role-based access control via Google Cloud Identity + Active Directory",
            "evidence": "Access Control Policy",
            "test_result": "Verified 10 sample users - all properly provisioned",
            "last_tested": now,
        },
        # AC.L2-3.1.1 - Second evidence
        {
            "control_id": "",
            "control_name": "",
            "status": "",
            "implementation": "",
            "evidence": "Q3 Access Review",
            "test_result": "",
            "last_tested": "",
        },
        # AC.L2-3.1.3
        {
            "control_id": "AC.L2-3.1.3",
            "control_name": AC_CONTROLS["AC.L2-3.1.3"]["name"],
            "status": "COMPLIANT",
            "implementation": "CUI isolated in dedicated VPC with DLP policies",
            "evidence": "Network Segmentation Diagram",
            "test_result": "Firewall rules validated, DLP blocking test successful",
            "last_tested": now,
        },
        # AC.L2-3.1.6 - First evidence
        {
            "control_id": "AC.L2-3.1.6",
            "control_name": AC_CONTROLS["AC.L2-3.1.6"]["name"],
            "status": "COMPLIANT",
            "implementation": "15-minute timeout enforced via GPO and Workspace policy",
            "evidence": "GPO Config",
            "test_result": "Tested 25 random workstations - all locked at 15 minutes",
            "last_tested": now,
        },
        # AC.L2-3.1.6 - Second evidence
        {
            "control_id": "",
            "control_name": "",
            "status": "",
            "implementation": "",
            "evidence": "Session Lock Test Results",
            "test_result": "",
            "last_tested": "",
        },
        # AC.L2-3.1.7
        {
            "control_id": "AC.L2-3.1.7",
            "control_name": AC_CONTROLS["AC.L2-3.1.7"]["name"],
            "status": "COMPLIANT",
            "implementation": "2-year minimum retention for all user identifiers",
            "evidence": "Identity Management Procedure",
            "test_result": "Verified naming convention prevents collisions",
            "last_tested": now,
        },
        # AC.L2-3.1.12
        {
            "control_id": "AC.L2-3.1.12",
            "control_name": AC_CONTROLS["AC.L2-3.1.12"]["name"],
            "status": "COMPLIANT",
            "implementation": "All remote access through VPN with MFA, sessions recorded",
            "evidence": "Remote Access Monitoring Report",
            "test_result": "SIEM alerts working, demonstrated anomaly detection",
            "last_tested": now,
        },
        # AC.L2-3.1.20 - First evidence
        {
            "control_id": "AC.L2-3.1.20",
            "control_name": AC_CONTROLS["AC.L2-3.1.20"]["name"],
            "status": "PARTIALLY_COMPLIANT",
            "implementation": "Continuous monitoring of 12 external connections",
            "evidence": "Connection Inventory",
            "test_result": "2 vendor attestations expired - access auto-restricted to read-only",
            "last_tested": now,
            "finding": "Vendor attestations expired for VEN-003 and VEN-007",
            "remediation": "Updated attestations expected by month-end",
        },
        # AC.L2-3.1.20 - Second evidence
        {
            "control_id": "",
            "control_name": "",
            "status": "",
            "implementation": "",
            "evidence": "Dashboard Screenshot",
            "test_result": "",
            "last_tested": "",
            "finding": "",
            "remediation": "",
        },
        # AC.L2-3.1.20 - Third evidence
        {
            "control_id": "",
            "control_name": "",
            "status": "",
            "implementation": "",
            "evidence": "Firewall Logs",
            "test_result": "",
            "last_tested": "",
            "finding": "",
            "remediation": "",
        },
        # AC.L2-3.1.20 - Fourth evidence
        {
            "control_id": "",
            "control_name": "",
            "status": "",
            "implementation": "",
            "evidence": "Alert Email",
            "test_result": "",
            "last_tested": "",
            "finding": "",
            "remediation": "",
        },
    ]
    return pd.DataFrame(evidence)


def generate_continuous_monitoring_data():
    """Generate continuous monitoring evidence for AC.L2-3.1.20"""
    vendors = [
        ("VEN-001", "Raytheon", "Valid", "2024-12-15"),
        ("VEN-002", "Lockheed", "Valid", "2024-11-30"),
        ("VEN-003", "Northrop", "EXPIRED", "2024-08-30"),
        ("VEN-004", "Boeing", "Valid", "2025-01-15"),
        ("VEN-007", "L3Harris", "EXPIRED", "2024-08-28"),
    ]
    now = (datetime.now() - timedelta(hours=2)).strftime("%Y-%m-%d %H:%M:%S")

    rows = []
    for ven_id, name, status, expiry in vendors:
        rows.append(
            {
                "connection_id": ven_id,
                "vendor_name": name,
                "connection_type": "VPN",
                "attestation_status": status,
                "expiry_date": expiry,
                "last_activity": now,
                "data_transferred_24h": f"{np.random.randint(50, 500)}MB",  # pyright: ignore[reportUndefinedVariable]
                "risk_score": "HIGH" if status == "EXPIRED" else "LOW",
                "access_level": "Read-Only" if status == "EXPIRED" else "Full",
                "monitoring_status": "Active",
                "alerts_sent": "Yes" if status == "EXPIRED" else "No",
            }
        )
    return pd.DataFrame(rows)


def generate_assessment_summary(evidence_df):
    """Generate executive summary of assessment"""
    # Count unique controls, not rows
    unique_controls = evidence_df[evidence_df['control_id'] != '']['control_id'].nunique()
    
    # Count status only for rows with control_id (not continuation rows)
    main_rows = evidence_df[evidence_df['control_id'] != '']
    compliant = len(main_rows[main_rows['status'] == 'COMPLIANT'])
    partial = len(main_rows[main_rows['status'] == 'PARTIALLY_COMPLIANT'])
    non_compliant = len(main_rows[main_rows['status'] == 'NON_COMPLIANT'])
    
    summary = {
        'Assessment Date': [MOCK_COMPANY['assessment_date']],
        'Company': [MOCK_COMPANY['name']],
        'Total AC Controls': [unique_controls],  # Should be 6, not 11
        'Fully Compliant': [compliant],  # Should be 5
        'Partially Compliant': [partial],  # Should be 1
        'Non-Compliant': [non_compliant],  # Should be 0
        'Compliance Rate': [f"{(compliant/unique_controls)*100:.1f}%"],
        'Critical Findings': [partial]
    }
    
    return pd.DataFrame(summary)


def _auto_fit_and_style(ws, zebra=True):
    """Apply header style, auto-width, filters, freeze panes, zebra rows, and borders."""
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Freeze header row & add filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Header row styling
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    # Auto-fit columns (cap width)
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            v_len = len(str(v)) if v is not None else 0
            max_len = max(max_len, v_len)
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    # Zebra striping + borders
    if zebra and ws.max_row >= 3:
        alt_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
        for r in range(2, ws.max_row + 1):
            if r % 2 == 0:
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.fill = alt_fill
                    cell.border = border
            else:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).border = border


def _color_cues(ws, sheet_name):
    """Contextual color cues per sheet."""
    # Evidence sheet: status in column 3
    if sheet_name == "AC_Controls_Evidence":
        status_col = 3
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=status_col).value
            if val == "COMPLIANT":
                ws.cell(row=r, column=status_col).fill = green
            elif val == "PARTIALLY_COMPLIANT":
                ws.cell(row=r, column=status_col).fill = yellow

    # External Monitoring: risk_score and attestation status
    if sheet_name == "External_Monitoring":
        # Find columns by header
        headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        risk_col = headers.get("risk_score")
        attest_col = headers.get("attestation_status")
        high = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # soft red
        warn = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # soft yellow

        if risk_col:
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=risk_col).value == "HIGH":
                    ws.cell(row=r, column=risk_col).fill = high

        if attest_col:
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(row=r, column=attest_col).value).upper() == "EXPIRED":
                    ws.cell(row=r, column=attest_col).fill = high
                elif str(ws.cell(row=r, column=attest_col).value).upper() == "VALID":
                    ws.cell(row=r, column=attest_col).fill = warn


def add_hyperlinks_to_evidence(worksheet):
    """Add hyperlinks to evidence column"""
    github_base = "https://github.com/securedbyjc/cmmc-ac-controls-demo/blob/main/evidence_samples/"
    
    # Define evidence links mapping - THIS WAS MISSING
    evidence_links = {
        2: [("AC_3.1.1_Access_Control_Policy.md", "Access Control Policy")],
        3: [("AC_3.1.1_Q3_Access_Review.csv", "Q3 Access Review")],
        4: [("AC_3.1.3_CUI_Network_Segmentation_Diagram.png", "Network Segmentation Diagram")],
        5: [("AC_3.1.6_Session_Lock_GPO.xml", "GPO Config")],
        6: [("AC_3.1.6_Session_Lock_Test_Results.md", "Session Lock Test Results")],
        7: [("AC_3.1.7_Identity_Management_Procedure.md", "Identity Management Procedure")],
        8: [("AC_3.1.12_Remote_Access_Monitoring_Report.md", "Remote Access Monitoring Report")],
        9: [("AC_3.1.20_External_Connection_Inventory.md", "Connection Inventory")],
        10: [("Screenshot_Dashboard_VEN003_Restricted.md", "Dashboard Screenshot")],
        11: [("Firewall_Log_VEN003_Restriction.log", "Firewall Logs")],
        12: [("Email_Alert_VEN003_Expiry.md", "Alert Email")],
    }
    
    # Add hyperlinks to evidence column (column 5)
    for row, links in evidence_links.items():
        if links:
            file, text = links[0]
            cell = worksheet.cell(row=row, column=5)
            cell.hyperlink = f"{github_base}{file}"
            cell.font = Font(color="0563C1", underline="single")


def create_excel_report(output_file: Path) -> Path:
    """Generate comprehensive Excel report with hyperlinks and styling"""
    # Generate data
    evidence = generate_ac_evidence()
    monitoring = generate_continuous_monitoring_data()
    summary = generate_assessment_summary(evidence)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Write all three sheets (index=False to keep it clean)
        evidence.to_excel(writer, sheet_name="AC_Controls_Evidence", index=False)
        summary.to_excel(writer, sheet_name="Executive_Summary", index=False)
        monitoring.to_excel(writer, sheet_name="External_Monitoring", index=False)

        # Access workbook and sheets
        wb = writer.book
        ws_evidence = writer.sheets["AC_Controls_Evidence"]
        ws_summary = writer.sheets["Executive_Summary"]
        ws_monitor = writer.sheets["External_Monitoring"]

        # Add hyperlinks in the evidence sheet
        add_hyperlinks_to_evidence(ws_evidence)

        # Apply consistent styling
        for name, ws in [
            ("AC_Controls_Evidence", ws_evidence),
            ("Executive_Summary", ws_summary),
            ("External_Monitoring", ws_monitor),
        ]:
            _auto_fit_and_style(ws, zebra=True)
            _color_cues(ws, name)

        # Executive summary: emphasize Compliance Rate cell
        # Locate "Compliance Rate" column and apply accent fill to its data cell
        headers = {ws_summary.cell(row=1, column=c).value: c for c in range(1, ws_summary.max_column + 1)}
        comp_col = headers.get("Compliance Rate")
        if comp_col:
            data_cell = ws_summary.cell(row=2, column=comp_col)
            data_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            data_cell.font = Font(bold=True)

    return output_file


def main():
    """Main execution"""
    parser = argparse.ArgumentParser(description="CMMC AC Controls Assessment Demo")
    parser.add_argument("--assessment", default="AC_CONTROLS", help="Assessment type")
    args = parser.parse_args()

    print(f"\n📋 Generating assessment for: {MOCK_COMPANY['name']}")
    print(f"👥 Employees: {MOCK_COMPANY['employees']} ({MOCK_COMPANY['cui_users']} with CUI access)")
    print(f"🔒 Focus: CMMC Level 2 Access Control Requirements\n")

    # Generate report
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = Path(f"output/cmmc_ac_assessment_{timestamp}.xlsx")
    output_file.parent.mkdir(exist_ok=True)

    print("⚙️  Collecting evidence from mock systems...")
    print("   ✓ Access control matrix")
    print("   ✓ Session policies")
    print("   ✓ Remote access logs")
    print("   ✓ External connection monitoring")

    report = create_excel_report(output_file)

    print(f"\n✅ Assessment complete!")
    print(f"📁 Report generated: {report}")
    print("\n📊 Results Summary:")
    print("   • 5 of 6 AC controls FULLY COMPLIANT")
    print("   • 1 control PARTIALLY COMPLIANT (AC.L2-3.1.20)")
    print("   • Finding: 2 vendor attestations expired (auto-restricted)")
    print("\n" + "=" * 60)
    print("This educational demo shows automated evidence generation.")
    print("For production use, implement real API integrations.")
    print("=" * 60)


if __name__ == "__main__":
    main()